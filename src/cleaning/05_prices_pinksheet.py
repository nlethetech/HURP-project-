#!/usr/bin/env python3
"""Tidy the World Bank Pink Sheet ANNUAL workbook into a long commodity-price table.

Purpose
-------
Reshape the World Bank Commodity Markets (Pink Sheet) ANNUAL historical-data
workbook into one tidy long table of annual commodity prices - nominal and real
(where the workbook supplies both) - plus the year-over-year log change of the
nominal price (`dlog_price`). All commodities are kept; the merge stage selects
the agricultural subset against the local crop mix. See docs/DATA_SOURCES.md,
"World Bank Pink Sheet".

Inputs
------
    data/raw/pink_sheet/CMO-Historical-Data-Annual.xlsx
        (produced by src/acquisition/05_download_pink_sheet.py)
        Relevant sheets:
          "Annual Prices (Nominal)" - commodity names on row 7, units on row 8,
              year in column A from row 9; values 1960-2025.
          "Annual Prices (Real)"    - same layout; the 69 commodities plus a
              trailing "MUV Index" (the manufactures-unit-value deflator).
        Missing values are encoded as the literal "..." (ellipsis) or "..".

Outputs
-------
    data/interim/prices_pinksheet.parquet
        One row per (commodity, year):
            commodity_code (str)   <- deterministic slug of the normalized name
            commodity_name (str)   <- footnote-stripped commodity label
            year           (int64)
            price          (float) <- annual NOMINAL price (NaN if unquoted)
            price_real     (float) <- annual REAL price where the workbook
                                      provides it, else NaN
            unit           (str)   <- price unit, e.g. "$/mt", "$/kg", "$/bbl"
            dlog_price     (float) <- log(price_t) - log(price_{t-1}) of the
                                      NOMINAL price, per commodity by year

Keys: this is a commodity x year table (no district_id / iso3 - the Pink Sheet
is a single global benchmark; spatial variation enters later via crop shares).

Determinism
-----------
Output is sorted by (commodity_code, year) and written without an index, so the
same workbook yields a byte-stable parquet.

Missing-value handling (constants below)
----------------------------------------
The workbook uses two textual sentinels for "not quoted": MISSING_SENTINELS.
Those cells become NaN. `dlog_price` is NaN wherever the current or prior-year
nominal price is missing or non-positive (log undefined), or at a commodity's
first year. The counts of sentinel cells and of NaN dlog values are logged.

Runtime
-------
A few seconds (reads one ~3 MiB workbook).

How to run
----------
    .venv/bin/python src/cleaning/05_prices_pinksheet.py

Idempotent: the output is rewritten in full each run from the immutable raw xlsx.
"""

from __future__ import annotations

import re
import sys
import unicodedata
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

# --- Constants ---------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parents[2]
RAW_XLSX = REPO_ROOT / "data" / "raw" / "pink_sheet" / "CMO-Historical-Data-Annual.xlsx"
OUT_PARQUET = REPO_ROOT / "data" / "interim" / "prices_pinksheet.parquet"

NOMINAL_SHEET = "Annual Prices (Nominal)"
REAL_SHEET = "Annual Prices (Real)"

# Sheet layout (1-based, as laid out by the World Bank): commodity names on
# row 7, units on row 8, annual data from row 9; column 1 is the year label.
NAME_ROW = 7
UNIT_ROW = 8
DATA_START_ROW = 9
YEAR_COL = 1  # 1-based; openpyxl row tuples are 0-based, so index 0.

# Textual "not quoted" sentinels used in the workbook cells.
MISSING_SENTINELS = {"…", "...", "..", "", "n.a.", "na", "-"}

# Plausibility floor/ceiling on the year label to reject stray header text.
MIN_YEAR = 1900
MAX_YEAR = 2100

FINAL_COLS = [
    "commodity_code",
    "commodity_name",
    "year",
    "price",
    "price_real",
    "unit",
    "dlog_price",
]
SORT_COLS = ["commodity_code", "year"]


def log(msg: str) -> None:
    print(msg, flush=True)


def normalize_name(raw: str) -> str:
    """Strip trailing asterisk footnote markers and collapse whitespace.

    The nominal and real price sheets differ only in asterisk footnote
    decoration for some series (e.g. nominal "Lamb **" vs real "Lamb",
    "Rubber, TSR20 **"); stripping trailing '*' runs makes the two sheets join
    on a stable commodity key. Only asterisks are removed - trailing digits are
    part of the commodity identity here (e.g. "Rice, Thai 5%", "Rubber, TSR20",
    "Rice, Thai A.1") and are preserved.
    """
    s = str(raw).strip()
    s = re.sub(r"\s*\*+\s*$", "", s)  # trailing '*'/'**' footnote markers only
    s = re.sub(r"\s+", " ", s).strip()
    return s


def slugify(name: str) -> str:
    """Deterministic ASCII slug used as commodity_code (e.g. 'rice_thai_5')."""
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def clean_unit(raw) -> str:
    """Strip the parentheses the workbook wraps around units, e.g. '($/mt)'."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
    return s.strip()


def to_number(value):
    """Return a float for numeric cells, NaN for the textual missing sentinels."""
    if value is None:
        return np.nan
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s in MISSING_SENTINELS:
        return np.nan
    # Any other non-numeric text is a parse surprise -> fail loudly upstream.
    return float(s)


def read_price_sheet(path: Path, sheet: str) -> tuple[pd.DataFrame, int]:
    """Read one annual price sheet into long form.

    Returns (long_df[commodity_name, code, unit, year, value], n_sentinel_cells).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet {sheet!r} not found in {path.name}; sheets={wb.sheetnames}")
    ws = wb[sheet]

    names_row = list(next(ws.iter_rows(min_row=NAME_ROW, max_row=NAME_ROW, values_only=True)))
    units_row = list(next(ws.iter_rows(min_row=UNIT_ROW, max_row=UNIT_ROW, values_only=True)))

    # Columns 2..N (0-based index 1..) are commodities; index 0 is the year label.
    cols = []  # (col_index, commodity_name, commodity_code, unit)
    for ci in range(1, len(names_row)):
        raw_name = names_row[ci]
        if raw_name is None or str(raw_name).strip() == "":
            continue  # trailing empty columns
        name = normalize_name(raw_name)
        if not name:
            continue
        unit = clean_unit(units_row[ci] if ci < len(units_row) else None)
        cols.append((ci, name, slugify(name), unit))

    n_sentinel = 0
    records = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, values_only=True):
        yr_cell = row[0]
        if yr_cell is None or str(yr_cell).strip() == "":
            continue
        try:
            year = int(str(yr_cell).strip())
        except ValueError:
            continue  # non-year stray row
        if not (MIN_YEAR <= year <= MAX_YEAR):
            continue
        for ci, name, code, unit in cols:
            cell = row[ci] if ci < len(row) else None
            val = to_number(cell)
            if np.isnan(val) and cell is not None and str(cell).strip() in MISSING_SENTINELS:
                n_sentinel += 1
            records.append(
                {
                    "commodity_name": name,
                    "commodity_code": code,
                    "unit": unit,
                    "year": year,
                    "value": val,
                }
            )
    wb.close()

    df = pd.DataFrame.from_records(records)
    return df, n_sentinel


def main() -> int:
    if not RAW_XLSX.exists():
        raise FileNotFoundError(
            f"Raw input missing: {RAW_XLSX}\n"
            "Run: .venv/bin/python src/acquisition/05_download_pink_sheet.py"
        )

    log(f"Reading {RAW_XLSX}")
    nominal, n_sent_nom = read_price_sheet(RAW_XLSX, NOMINAL_SHEET)
    real, n_sent_real = read_price_sheet(RAW_XLSX, REAL_SHEET)
    log(f"  nominal: {nominal['commodity_code'].nunique()} commodities, "
        f"{nominal['year'].nunique()} years, {len(nominal):,} cells "
        f"({n_sent_nom} missing-sentinel)")
    log(f"  real:    {real['commodity_code'].nunique()} commodities, "
        f"{real['year'].nunique()} years, {len(real):,} cells "
        f"({n_sent_real} missing-sentinel)")

    # A commodity must carry a single unit on each sheet; guard against drift.
    for label, frame in (("nominal", nominal), ("real", real)):
        u = frame.groupby("commodity_code")["unit"].nunique()
        bad = u[u > 1]
        if len(bad):
            raise ValueError(f"{label} sheet: commodities with >1 unit: {list(bad.index)}")

    # --- Assemble the (commodity_code, year) spine from the UNION of sheets ---
    nom = nominal.rename(columns={"value": "price"})
    rea = real.rename(columns={"value": "price_real"})[
        ["commodity_code", "year", "price_real"]
    ]

    # commodity_name + unit come from the nominal sheet; fall back to the real
    # sheet for commodities present only there (e.g. the MUV Index deflator).
    meta_nom = nom[["commodity_code", "commodity_name", "unit"]].drop_duplicates(
        "commodity_code"
    )
    meta_real = real[["commodity_code", "commodity_name", "unit"]].drop_duplicates(
        "commodity_code"
    )
    meta = pd.concat([meta_nom, meta_real]).drop_duplicates("commodity_code", keep="first")

    # Outer-merge nominal and real on the (commodity_code, year) key.
    df = pd.merge(
        nom[["commodity_code", "year", "price"]],
        rea,
        on=["commodity_code", "year"],
        how="outer",
        validate="one_to_one",
    )
    df = df.merge(meta, on="commodity_code", how="left", validate="many_to_one")

    if df["commodity_name"].isna().any() or df["unit"].isna().any():
        missing = df.loc[df["commodity_name"].isna() | df["unit"].isna(), "commodity_code"].unique()
        raise ValueError(f"Missing name/unit metadata for: {list(missing)}")

    # --- dlog_price: year-over-year log change of the NOMINAL price ----------
    df = df.sort_values(SORT_COLS, kind="mergesort").reset_index(drop=True)
    # log() is only defined for strictly positive prices.
    safe_price = df["price"].where(df["price"] > 0)
    log_price = np.log(safe_price)
    df["dlog_price"] = log_price.groupby(df["commodity_code"]).diff()

    n_real_commodities = df.loc[df["price_real"].notna(), "commodity_code"].nunique()
    n_dlog = int(df["dlog_price"].notna().sum())
    n_dlog_nan = int(df["dlog_price"].isna().sum())

    # --- Finalize types, order, sort -----------------------------------------
    df["year"] = df["year"].astype("int64")
    df["commodity_code"] = df["commodity_code"].astype(str)
    df["commodity_name"] = df["commodity_name"].astype(str)
    df["unit"] = df["unit"].astype(str)
    df["price"] = df["price"].astype(float)
    df["price_real"] = df["price_real"].astype(float)
    df["dlog_price"] = df["dlog_price"].astype(float)

    df = df[FINAL_COLS].sort_values(SORT_COLS, kind="mergesort").reset_index(drop=True)

    # Hard invariant: (commodity_code, year) is the unique primary key.
    if df.duplicated(["commodity_code", "year"]).any():
        raise AssertionError("(commodity_code, year) is not unique; aborting.")

    OUT_PARQUET.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(OUT_PARQUET, index=False)
    log(f"Wrote {OUT_PARQUET} ({len(df):,} rows)")

    # --- Summary -------------------------------------------------------------
    log("")
    log("=== PINK SHEET (ANNUAL) SUMMARY ===")
    log(f"  commodities:            {df['commodity_code'].nunique()}")
    log(f"  ... with a real price:  {n_real_commodities}")
    log(f"  year range:             {int(df['year'].min())}-{int(df['year'].max())}")
    log(f"  rows:                   {len(df):,}")
    log(f"  nominal prices present: {int(df['price'].notna().sum()):,}")
    log(f"  real prices present:    {int(df['price_real'].notna().sum()):,}")
    log(f"  dlog_price present:     {n_dlog:,} (NaN {n_dlog_nan:,})")
    log("OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
