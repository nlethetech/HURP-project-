#!/usr/bin/env python3
"""Parse the World Bank OGHIST workbook into a tidy income-class country-year table.

Purpose
-------
Turn the "Country Analytical History" sheet of the World Bank OGHIST workbook
(raw, downloaded by src/acquisition/03_download_wb_income.py) into a long,
deterministic country x fiscal-year income-classification table for the panel
(see docs/DATA_SOURCES.md, "World Bank Historical Country Income
Classifications (OGHIST)").

The sheet is a wide matrix: one row per economy, one column per World Bank
fiscal year FY89..FY26 (38 columns). Two header rows carry the alignment:
    row 5  "Bank's fiscal year:"      -> FYnn   (effectivity year, set July 1)
    row 6  "Data for calendar year :" -> 1987..2024 (the Atlas-GNI calendar year)
Cell values are the income class: L / LM / UM / H, the explicit "not classified"
code "..", a blank (no classification on record), or "LM*" (the pre-unification
Yemen footnote, classified Lower-middle).

Inputs
------
    data/raw/wb_income_oghist/OGHIST.xlsx
        (produced by src/acquisition/03_download_wb_income.py; sheet
         "Country Analytical History"; col A = WB economy code, col B = name)

Output
------
    data/interim/income_class.parquet
        One row per (economy, fiscal year) that carries a code (blank cells are
        dropped, see below):
            iso3         (str)  country/economy key  [see WB_CODE_TO_ISO3]
            fiscal_year  (int)  1989..2026           [from the FYnn header]
            data_year    (int)  1987..2024           [Atlas-GNI calendar year]
            income_group (str)  L | LM | UM | H | NA  [normalized; NA == ".."]
            raw_code     (str)  verbatim cell value   [incl. ".." and "LM*"]
        Sorted deterministically by (iso3, fiscal_year).

Encoding of cell values (registry "Gotchas" 4)
----------------------------------------------
  * "L" / "LM" / "UM" / "H"  -> income_group as-is, raw_code as-is.
  * "LM*"  -> income_group "LM" (pre-unification Yemen footnote, row 231);
              raw_code preserved as "LM*" so the flag is not lost.
  * ".."   -> economy on record but NOT classified that year; income_group set
              to the sentinel NOT_CLASSIFIED ("NA"); raw_code "..". Row KEPT
              (the economy existed and was explicitly unclassified).
  * blank  -> no classification on record (e.g. Venezuela FY22+, or a former
              entity after it dissolved). Row DROPPED: there is nothing to
              broadcast. Count logged (DROP_BLANK).
Do NOT coerce ".." or blank to "L"; both are non-low.

WB economy code -> ISO3 mapping (explicit, registry "Gotchas" 6)
---------------------------------------------------------------
Col A codes are 3-letter ISO3 for every active economy in this workbook EXCEPT
two, handled by the explicit WB_CODE_TO_ISO3 dict below:
  * "CHI"  -> Channel Islands: a World Bank reporting aggregate (Jersey +
             Guernsey), which has no single standard ISO3. Kept verbatim as
             "CHI" and flagged here; it will simply not join to a spine unit
             (the boundary spine splits these as JEY/GGY, absent from this file).
  * "YUGf" -> Yugoslavia (former), a non-ISO World Bank code for the pre-1992
             SFRY. Kept verbatim as "YUGf" (its successor states classify
             separately from FY93 on); a country-succession crosswalk, not this
             table, is the place to reconcile 1987-1991 successor mapping.
All other former-entity codes in the workbook (CSK Czechoslovakia, SUN USSR,
YUG Serbia and Montenegro, ANT Netherlands Antilles, MYT Mayotte) are already
valid ISO3 / WB codes and are retained verbatim; they are documented former or
territorial entities that will not match a modern boundary spine, which is
expected and handled at merge time, not here.

Row-level handling (logged; thresholds/constants at top of script)
------------------------------------------------------------------
  * Only rows whose col A is a non-empty economy code are read (this skips the
    title rows 1-3, the header/threshold rows 4-10, the blank row 11, the
    inter-block blank/footnote rows 230-232, and the July-1 note at row 240 --
    "read until blank" would truncate the 6 former entities, so we filter on a
    present code instead, per registry "Gotchas" 5).
  * DROP_BLANK: melted (economy, FY) cells with no value are dropped.

Runtime
-------
A few seconds (one small workbook).

How to run
----------
    .venv/bin/python src/cleaning/03_income_class.py

Idempotent: the output parquet is rewritten in full each run from the immutable
raw workbook, with a fixed sort order, so re-runs are byte-stable.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

# --- Constants ---------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parents[2]
RAW_XLSX = REPO_ROOT / "data" / "raw" / "wb_income_oghist" / "OGHIST.xlsx"
OUT_PARQUET = REPO_ROOT / "data" / "interim" / "income_class.parquet"

SHEET = "Country Analytical History"

# Header / data geometry of the sheet (1-based rows; verified against the source).
ROW_FISCAL_YEAR = 5      # "Bank's fiscal year:" -> FY89..FY26
ROW_DATA_YEAR = 6        # "Data for calendar year :" -> 1987..2024
DATA_FIRST_ROW = 12      # first economy row (AFG)
DATA_LAST_ROW = 238      # last economy row (YUGf, former-entities block end)
COL_CODE = 1             # col A: WB economy code (1-based)
COL_NAME = 2             # col B: economy name (1-based)
FY_FIRST_COL = 3         # col C: first FY column (1-based)

# Expected coverage (registry): FY89..FY26 == 38 fiscal years; data years
# 1987..2024. Asserted after parsing so a layout change fails loudly.
EXPECTED_N_FY = 38
EXPECTED_FY_RANGE = (1989, 2026)
EXPECTED_DATA_YEAR_RANGE = (1987, 2024)

# Cell-value vocabulary.
VALID_GROUPS = {"L", "LM", "UM", "H"}
NOT_CLASSIFIED_CODE = ".."        # economy on record, explicitly unclassified
NOT_CLASSIFIED_GROUP = "NA"       # sentinel income_group for ".."
YEMEN_PREUNIF_CODE = "LM*"        # pre-unification Yemen footnote -> LM

# Explicit, documented World Bank economy-code -> ISO3 overrides (see docstring).
# Every other col-A code is a valid ISO3 / WB code retained verbatim.
WB_CODE_TO_ISO3 = {
    "CHI": "CHI",    # Channel Islands: WB aggregate, no single ISO3; kept as-is
    "YUGf": "YUGf",  # Yugoslavia (former): non-ISO WB code; kept as-is
}

FINAL_COLS = ["iso3", "fiscal_year", "data_year", "income_group", "raw_code"]
SORT_COLS = ["iso3", "fiscal_year"]


def log(msg: str) -> None:
    print(msg, flush=True)


def fy_label_to_year(label: str) -> int:
    """'FY89' -> 1989, 'FY00' -> 2000, 'FY26' -> 2026 (two-digit FY century pivot).

    The OGHIST FY columns run FY89..FY26: a two-digit year >= 89 is 19xx, else
    20xx. This is exact for the documented FY89..FY26 span.
    """
    m = re.fullmatch(r"FY(\d{2})", label.strip())
    if not m:
        raise ValueError(f"Unexpected fiscal-year header {label!r}; expected 'FYnn'.")
    yy = int(m.group(1))
    return 1900 + yy if yy >= 89 else 2000 + yy


def main() -> int:
    if not RAW_XLSX.exists():
        raise FileNotFoundError(
            f"Raw input missing: {RAW_XLSX}\n"
            "Run: .venv/bin/python src/acquisition/03_download_wb_income.py"
        )

    log(f"Reading {RAW_XLSX} :: sheet {SHEET!r}")
    wb = openpyxl.load_workbook(RAW_XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet {SHEET!r} not present; found {wb.sheetnames}. Source changed."
        )
    ws = wb[SHEET]

    # --- Header rows: FY labels and data years -------------------------------
    fy_row = [c.value for c in next(ws.iter_rows(min_row=ROW_FISCAL_YEAR,
                                                 max_row=ROW_FISCAL_YEAR))]
    dy_row = [c.value for c in next(ws.iter_rows(min_row=ROW_DATA_YEAR,
                                                 max_row=ROW_DATA_YEAR))]

    # FY columns are contiguous from FY_FIRST_COL while the FY header is non-empty.
    fy_cols: list[tuple[int, int, int]] = []  # (col_index_0based, fiscal_year, data_year)
    for j in range(FY_FIRST_COL - 1, len(fy_row)):
        label = fy_row[j]
        if label is None or str(label).strip() == "":
            break
        fiscal_year = fy_label_to_year(str(label))
        data_year = dy_row[j]
        if not isinstance(data_year, int):
            raise ValueError(
                f"Data-for-calendar-year header at column {j + 1} is "
                f"{data_year!r}; expected an integer aligned with {label!r}."
            )
        fy_cols.append((j, fiscal_year, data_year))

    log(f"  detected {len(fy_cols)} fiscal-year columns "
        f"({fy_cols[0][1]}..{fy_cols[-1][1]}); "
        f"data years {fy_cols[0][2]}..{fy_cols[-1][2]}")

    # Hard guards: coverage must match the registry exactly.
    if len(fy_cols) != EXPECTED_N_FY:
        raise ValueError(
            f"Expected {EXPECTED_N_FY} FY columns, found {len(fy_cols)}; aborting."
        )
    fy_years = [fc[1] for fc in fy_cols]
    if (min(fy_years), max(fy_years)) != EXPECTED_FY_RANGE:
        raise ValueError(
            f"Fiscal-year range {min(fy_years)}..{max(fy_years)} != "
            f"{EXPECTED_FY_RANGE}; aborting."
        )
    dy_years = [fc[2] for fc in fy_cols]
    if (min(dy_years), max(dy_years)) != EXPECTED_DATA_YEAR_RANGE:
        raise ValueError(
            f"Data-year range {min(dy_years)}..{max(dy_years)} != "
            f"{EXPECTED_DATA_YEAR_RANGE}; aborting."
        )

    # --- Economy rows: melt the FY matrix to long ----------------------------
    records: list[dict] = []
    n_economy_rows = 0
    n_blank_cells = 0
    for row in ws.iter_rows(min_row=DATA_FIRST_ROW, max_row=DATA_LAST_ROW):
        values = [c.value for c in row]
        code = values[COL_CODE - 1]
        # Skip non-economy rows (blank rows, footnotes, the July-1 note): they
        # have no col-A code. Filtering on a present code (not "until blank")
        # preserves the 6 former-entity rows after the inter-block blanks.
        if code is None or str(code).strip() == "":
            continue
        code = str(code).strip()
        n_economy_rows += 1

        iso3 = WB_CODE_TO_ISO3.get(code, code)

        for j, fiscal_year, data_year in fy_cols:
            raw = values[j]
            if raw is None or str(raw).strip() == "":
                n_blank_cells += 1
                continue  # DROP_BLANK: no classification on record
            raw = str(raw).strip()
            records.append(
                {
                    "iso3": iso3,
                    "fiscal_year": fiscal_year,
                    "data_year": data_year,
                    "raw_code": raw,
                    "_code_in": code,
                }
            )

    log(f"  economy rows read:        {n_economy_rows}")
    log(f"  blank cells dropped:      {n_blank_cells} (DROP_BLANK; no class on record)")

    df = pd.DataFrame.from_records(records)
    if df.empty:
        raise ValueError("No income-class records parsed; aborting.")

    # --- Normalize income_group from raw_code --------------------------------
    def normalize(raw: str) -> str:
        if raw in VALID_GROUPS:
            return raw
        if raw == YEMEN_PREUNIF_CODE:
            return "LM"               # pre-unification Yemen -> Lower-middle
        if raw == NOT_CLASSIFIED_CODE:
            return NOT_CLASSIFIED_GROUP
        raise ValueError(
            f"Unrecognized income code {raw!r} (iso3 not given here); "
            "vocabulary changed -- update the cleaner. Aborting."
        )

    df["income_group"] = df["raw_code"].map(normalize)

    # Log the encoding-handling counts (registry gotcha 4 surfaces).
    n_dotdot = int((df["raw_code"] == NOT_CLASSIFIED_CODE).sum())
    n_yemen = int((df["raw_code"] == YEMEN_PREUNIF_CODE).sum())
    log(f"  '..' (not classified) kept: {n_dotdot} (income_group={NOT_CLASSIFIED_GROUP!r})")
    log(f"  'LM*' (pre-unif. Yemen):    {n_yemen} (income_group='LM', raw_code preserved)")

    # Surface the explicit WB-code overrides actually exercised.
    overridden = sorted(
        df.loc[df["_code_in"] != df["iso3"], "_code_in"].unique().tolist()
    )
    if overridden:
        log(f"  WB-code overrides applied:  {overridden} "
            "(see WB_CODE_TO_ISO3; documented)")

    # --- Final shape, types, deterministic order -----------------------------
    df["fiscal_year"] = df["fiscal_year"].astype("int64")
    df["data_year"] = df["data_year"].astype("int64")
    df["iso3"] = df["iso3"].astype(str)
    df["income_group"] = df["income_group"].astype(str)
    df["raw_code"] = df["raw_code"].astype(str)

    df = df[FINAL_COLS].copy()

    # Invariant: (iso3, fiscal_year) is unique (one class per economy per FY).
    dup = int(df.duplicated(subset=["iso3", "fiscal_year"]).sum())
    if dup:
        raise AssertionError(
            f"{dup} duplicate (iso3, fiscal_year) rows; key is not unique. Aborting."
        )

    df = df.sort_values(SORT_COLS, kind="mergesort").reset_index(drop=True)

    # --- Write output --------------------------------------------------------
    OUT_PARQUET.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(OUT_PARQUET, index=False)
    log(f"Wrote {OUT_PARQUET} ({len(df):,} rows)")

    # --- Summary -------------------------------------------------------------
    log("")
    log("=== INCOME CLASS SUMMARY ===")
    log(f"  rows:                {len(df):,}")
    log(f"  distinct economies:  {df['iso3'].nunique()}")
    log(f"  fiscal_year range:   {df['fiscal_year'].min()}..{df['fiscal_year'].max()}")
    log(f"  data_year range:     {df['data_year'].min()}..{df['data_year'].max()}")
    log("  income_group counts (all years):")
    for grp, cnt in df["income_group"].value_counts().sort_index().items():
        log(f"      {grp}: {cnt:,}")
    log("OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
